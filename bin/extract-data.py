# /usr/bin/env python

import base64
import coloredlogs
from datetime import datetime
import json
import hashlib
import logging as log
import os
import os.path
import pprint
from shutil import copyfile
import subprocess
import itertools
import sys
from types import SimpleNamespace
from openpyxl import load_workbook

coloredlogs.install()
pp = pprint.PrettyPrinter(compact=True)

if os.getenv("LOG") == "DEBUG":
    log.basicConfig(level=log.DEBUG)
else:
    log.basicConfig(level=log.WARN)

WORDS = [
    "welcome",
    "hello",
    "goodbye/I'll see you later",
    "where are you going?",
    "I'm going home",
    "no",
    "yes",
    "what is your name?",
    "come here",
    "let's go",
    "fire",
    "firewood",
    "smoke",
    "ashes",
    "hot weather",
    "cold weather",
    "wind",
    "rain",
    "sun",
    "moon",
    "star",
    "Southern Cross",
    "Seven Sisters",
    "sky",
    "Milky Way",
    "cloud",
    "hand",
    "elbow",
    "eye",
    "nose",
    "mouth",
    "thigh",
    "foot",
    "head",
    "knee",
    "shoulder",
    "ear",
    "tree",
    "fish",
    "bird",
    "goanna",
    "grey kangaroo",
    "emu",
    "magpie",
    "water",
    "dog",
    "here",
    "there",
    "today",
    "tomorrow",
    "mother",
    "father",
    "brother",
    "sister",
    "child",
    "bone",
    "money",
]


class SheetVerifier:
    def __init__(self, rows, sheet_name):
        self.language = sheet_name.replace("/srv/data/", "").split("/")[0]
        self.sheet_name = sheet_name.replace("/srv/data/", "")
        self.ok = True
        self.errors = []
        self.rows = rows
        # self.rows = [row for row in self.sheet.rows if row[0].value is not None]
        # for row in self.rows:
        #     print(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value)

    def verify(self):
        if len(self.rows) != 65:
            self.ok = False
            self.errors.append(
                {
                    "type": "Bad spreadsheet",
                    "level": "error",
                    "msg": f"'{self.sheet_name}' isn't exactly 65 rows - is it correct?",
                }
            )
        self.check(0, 0, "Language name")
        self.check(0, 1)
        self.check(0, 2)
        self.check(1, 0, "AIATSIS code")
        self.check(1, 1)
        self.check(2, 0, "Speaker's name")
        self.check(2, 1)
        self.check(2, 2)
        self.check(3, 0, "Other people who helped to get the list produced")
        self.check(3, 1)
        self.check(6, 0, "Date received")
        self.check(6, 1)
        j = 0
        for i in range(8, 65):
            row = self.rows
            if not row[i][1].value:
                self.errors.append(
                    {
                        "type": "Sheet verification: No translation for word",
                        "level": "warning",
                        "language": self.language,
                        "msg": f"No translation for word '{row[i][0].value}'.",
                    }
                )
            if not row[i][2].value:
                self.errors.append(
                    {
                        "type": "Sheet verification: No media file for word",
                        "level": "warning",
                        "language": self.language,
                        "msg": f"No media file for word '{row[i][0].value}'.",
                    }
                )

            if row[i][0].value != WORDS[j]:
                self.errors.append(
                    {
                        "type": "Sheet verification: English word has been changed",
                        "level": "warning",
                        "language": self.language,
                        "msg": f"Expected '{WORDS[j]}' but got '{row[i][0].value}'.",
                    }
                )

            j += 1
        return self.errors

    def check(self, rowNum, colNum, value=None):
        row = self.rows[rowNum]
        if value and row[colNum].value.strip() != value:
            self.ok = False
            self.errors.append(
                {
                    "type": "Sheet verification: incorrect data",
                    "level": "error",
                    "language": self.language,
                    "msg": f"Unexpected value in row: {rowNum}, column: {colNum}. Expected: {value}, Got: {row[colNum]}",
                }
            )
        elif not row[colNum]:
            self.errors.append(
                {
                    "type": "Sheet verification: missing data",
                    "level": "warning",
                    "language": self.language,
                    "msg": f"Empty cell found at row: {rowNum+1}, column: {colNum+1}.",
                }
            )


class DataExtractor:
    def __init__(self):
        self.aiatsis_geographies_by_name = {}
        self.aiatsis_geographies_by_code = {}
        self.gambay_geographies = {}
        self.errors = []
        self.data = {}
        self.words = {}
        self.languages = {}
        # self.gambay_additions = []
        # remember that these paths are relative to the volume
        #  mountpoints inside the container
        self.data_path = "/srv/data"
        self.repository = "/srv/dist/repository"
        self.aiatsis_geography_file = f"{self.data_path}/AIATSIS-geography.xlsx"
        self.gambay_geography_file = f"{self.data_path}/gambay-languages.geojson"

    def extract(self):
        self.extract_aiatsis_geographies()
        self.extract_gambay_geographies()
        self.map_gambay_and_aiatsis_geographies()
        self.apply_aiatsis_overrides()
        self.convert_to_geojson_features()
        self.remove_languages_without_geo_data()
        data = self.locate_languages_with_data()
        # # pp.pprint(data)
        data = self.extract_language_data(data)
        for item in data:
            self.data[item["code"]]["properties"]["language"]["name"] = item[
                "language"
            ]["name"]
            self.data[item["code"]]["properties"]["name"] = item["language"]["name"]
        self.build_repository(data)
        self.write_master_indices()
        # pp.pprint(self.errors)

    def extract_aiatsis_geographies(self):
        def parse_row(row):
            return {
                "code": row[0].value,
                "name": row[1].value,
                "lat": row[3].value,
                "lng": row[4].value,
                "override": row[7].value,
            }

        log.info("Extracting AIATSIS geography data")
        data = []
        wb = load_workbook(filename=self.aiatsis_geography_file)
        sheet = wb.worksheets[0]
        for row in sheet.rows:
            row = parse_row(row)
            self.aiatsis_geographies_by_code[row["code"]] = row
            self.aiatsis_geographies_by_name[row["name"]] = row
        data = itertools.groupby(data, key=lambda i: i["code"])
        for key, group in data:
            if len(list(group)) > 1:
                print(f"{key} has more than entry in the AIATSIS geography file")
                self.errors.append(
                    {
                        "type": "AIATSIS code found twice",
                        "level": "error",
                        "msg": f"The AIATSIS sheet has more than one entry with the code {key}",
                    }
                )

    def extract_gambay_geographies(self):
        def parse(item):
            return {
                "code": item["properties"]["austlang"]
                if "austlang" in item["properties"]
                else "",
                "name": item["properties"]["name"],
                "lat": item["geometry"]["coordinates"][1],
                "lng": item["geometry"]["coordinates"][0],
            }

        log.info("Extracting Gambay geography data")
        with open(self.gambay_geography_file, "r") as f:
            gambay_data = json.load(f)

        for language in gambay_data["features"]:
            item = parse(language)
            if not (item["code"]):
                self.errors.append(
                    {
                        "type": "Gambay entry missing Austlang code",
                        "level": "error",
                        "msg": f"Gambay item '{item['name']}'' is missing an Austlang Code.",
                    }
                )

            if item["code"].upper() != item["code"]:
                self.errors.append(
                    {
                        "type": "Gambay code lowercased",
                        "level": "error",
                        "msg": f"Gambay code for {item['name']} is lowercase: '{item['code']}'. Should be {item['code'].upper()}",
                    }
                )
                item["code"] = item["code"].upper()
            # pp.pprint(item)
            self.gambay_geographies[item["name"]] = item
        # for item in self.gambay_geographies.items():
        #     pp.pprint(item)

    def map_gambay_and_aiatsis_geographies(self):
        log.info("Mapping Gambay and AIATSIS geographies")
        for (name, item) in self.gambay_geographies.items():
            if "code" in item and not "#" in item["code"]:
                item["source"] = "Gambay"
                self.data[item["code"]] = item
            elif item["name"] in self.aiatsis_geographies_by_name:
                item = self.aiatsis_geographies[item["name"]]
                item["source"] = "Austlang"
                self.data[item["code"]] = item
            else:
                self.errors.append(
                    {
                        "type": "Item missing code in Gambay and Gambay name not found in Austlang",
                        "level": "error",
                        "msg": f"Item with name {item['name']} has no code in Gambay and this name is not found in Austlang",
                    }
                )

    def apply_aiatsis_overrides(self):
        log.info("Applying AIATSIS overrides")
        for name, item in self.aiatsis_geographies_by_code.items():
            item["source"] = "Austlang"
            if item["override"]:
                self.data[item["code"]] = item
            if item["code"] not in self.data:
                self.data[item["code"]] = item

    def convert_to_geojson_features(self):
        log.info("Converting the data to GeoJSON features")
        for name, item in self.data.items():
            self.data[item["code"]] = {
                "type": "Feature",
                "geometry": {
                    "coordinates": [item["lng"], item["lat"]],
                    "type": "Point",
                },
                "properties": {
                    "code": item["code"],
                    "name": item["name"],
                    "source": item["source"],
                    "selected": False,
                },
            }

    def remove_languages_without_geo_data(self):
        log.info("Remove languages without Geo data")
        data = {}
        for (key, item) in self.data.items():
            geo = item["geometry"]["coordinates"]
            if geo[0] != "" and geo[1] != "":
                data[key] = item
        self.data = data

    def locate_languages_with_data(self):
        data = []
        for root, dirs, files in os.walk(self.data_path):
            if root == "/srv/data":
                continue

            log.info(f"Processing: {root}")
            sheet = []
            for file in files:
                if "xlsx" in file and not "~$" in file:
                    sheet.append(file)

            if len(sheet) == 0:
                self.errors.append(
                    {
                        "type": "No spreadsheet",
                        "level": "error",
                        "msg": f"No spreadsheet in '{root}'. Skipping this folder.",
                    }
                )
                continue

            if len(sheet) > 1:
                self.errors.append(
                    {
                        "type": "Multiple spreadsheets",
                        "level": "error",
                        "msg": f"Found more than one data spreadsheet in folder '{root}'. Skipping this folder.",
                    }
                )
                continue

            data.append(
                {
                    "root": root,
                    "name": os.path.join(root, sheet[0]).replace("/srv/data/", ""),
                    "sheet": os.path.join(root, sheet[0]),
                }
            )
        return data

    def extract_language_data(self, data):
        def get(item):
            try:
                return item.value.strip() if item.value is not None else ""
            except:
                return item.value if item.value is not None else ""

        def parse_row(row):
            data = {"english": get(row[0]), "indigenous": get(row[1])}
            if ".mov" in get(row[2]):
                data["video_file"] = get(row[2])
            elif ".wav" in get(row[2]):
                data["audio_file"] = get(row[2])

            try:
                if get(row[3]):
                    data["english_alternate"] = get(row[3])
            except:
                pass
            return data

        itemsWithData = []
        for item in data:
            sheetname = item["name"]
            worksheet = item["sheet"]
            root = item["root"]

            log.info(f"Processing sheet: {sheetname}")
            try:
                wb = load_workbook(filename=worksheet)
            except:
                self.errors.append(
                    {
                        "type": "Critical errors in spreadsheet",
                        "level": "Error",
                        "msg": f"There were errors in the spreadsheet {sheetname} so this language has been skipped",
                    }
                )
                continue

            sheet = wb.worksheets[0]

            i = 0
            rows = []
            for row in sheet.rows:
                if i < 65:
                    rows.append(row)
                i += 1

            v = SheetVerifier(rows, worksheet)
            errors = v.verify()
            self.errors.extend(errors)
            if not v.ok:
                self.errors.append(
                    {
                        "type": "Critical errors in spreadsheet",
                        "level": "Error",
                        "msg": f"There were errors in the spreadsheet {sheetname} so this language has been skipped",
                    }
                )
                continue

            log.info(f"Extracting data: {sheetname}")
            sheet = {
                "language": {
                    "name": get(rows[0][1]),
                    "audio_file": os.path.join(root, get(rows[0][2]))
                    if get(rows[0][2])
                    else "",
                },
                "date_received": get(rows[6][1]).strftime("%Y%m%d")
                if isinstance(get(rows[6][1]), datetime)
                else get(rows[6][1]),
                "code": get(rows[1][1]),
                "weblink": get(rows[5][1]),
                "words": [],
                "speaker": {
                    "name": get(rows[2][1]),
                    "audio_file": os.path.join(root, get(rows[2][2]))
                    if get(rows[2][2])
                    else "",
                },
                "thankyou": get(rows[3][1]),
            }
            if sheet["code"] not in self.data:
                self.errors.append(
                    {
                        "type": "Code not found",
                        "level": "error",
                        "msg": f"'{sheetname}' has code '{sheet['code']}' but that code is not in Gambay or AIATSIS geo data.",
                    }
                )

            for r in range(8, len(rows)):
                data = parse_row(rows[r])
                if "audio_file" in data:
                    data["audio_file"] = os.path.join(root, data["audio_file"])
                elif "video_file" in data:
                    data["video_file"] = os.path.join(root, data["video_file"])

                if "audio_file" in data or "video_file" in data:
                    sheet["words"].append(data)

            self.data[sheet["code"]]["properties"] = {
                **sheet,
                **self.data[sheet["code"]]["properties"],
            }
            itemsWithData.append(sheet)
        return itemsWithData

    def build_repository(self, data):
        def get_target_name(path, file, ext):
            return os.path.join(path, os.path.splitext(os.path.basename(file))[0]) + ext

        def transcode(item, target, format):
            if os.environ["UPDATE_ALL"] == "true" or not os.path.exists(target):
                log.debug(f"Transcoding {item} to {format}")
                subprocess.run(
                    [
                        "ffmpeg",
                        "-hide_banner",
                        "-loglevel",
                        "panic",
                        "-y",
                        "-i",
                        item,
                        target,
                    ]
                )

        def transcode_and_copy_to_repository(item, item_path):
            if "video_file" in item:
                video_file = item["video_file"]
                if not video_file:
                    del item["video_file"]
                    item["video"] = []
                    return item

                if not os.path.exists(video_file):
                    print(item)
                    self.errors.append(
                        {
                            "type": "Video file not found",
                            "level": "error",
                            "msg": f"'{video_file}' specified in sheet but file not found",
                        }
                    )
                    del item["video_file"]
                    item["video"] = []
                    return item

                transcode(
                    video_file, get_target_name(item_path, video_file, ".webm"), "webm"
                )
                transcode(
                    video_file, get_target_name(item_path, video_file, ".mp4"), "mp4"
                )
                video_files = [
                    get_target_name(item_path, video_file, ".webm").replace(
                        "/srv/dist", ""
                    ),
                    get_target_name(item_path, video_file, ".mp4").replace(
                        "/srv/dist", ""
                    ),
                ]
                copyfile(
                    video_file, os.path.join(item_path, os.path.basename(video_file))
                )
                video_files.append(
                    os.path.join(item_path, os.path.basename(video_file)).replace(
                        "/srv/dist", ""
                    )
                )

                item["video"] = video_files
                del item["video_file"]
                return item

            if "audio_file" in item:
                audio_file = item["audio_file"]
                if not audio_file:
                    del item["audio_file"]
                    item["audio"] = []
                    return item

                if not os.path.exists(audio_file):
                    self.errors.append(
                        {
                            "type": "Audio file missing",
                            "level": "error",
                            "msg": f"'{audio_file}' specified in sheet but file not found",
                        }
                    )
                    del item["audio_file"]
                    item["audio"] = []
                    return item

                transcode(
                    audio_file, get_target_name(item_path, audio_file, ".webm"), "webm"
                )
                transcode(
                    audio_file, get_target_name(item_path, audio_file, ".mp3"), "mp3"
                )
                audio_files = [
                    get_target_name(item_path, audio_file, ".webm").replace(
                        "/srv/dist", ""
                    ),
                    get_target_name(item_path, audio_file, ".mp3").replace(
                        "/srv/dist", ""
                    ),
                ]
                if "wav" in audio_file:
                    copyfile(
                        audio_file,
                        os.path.join(item_path, os.path.basename(audio_file)),
                    )
                    audio_files.append(
                        os.path.join(item_path, os.path.basename(audio_file)).replace(
                            "/srv/dist", ""
                        )
                    )

                item["audio"] = audio_files
                del item["audio_file"]
                return item

        def push_to_words(word, item):
            # item = {**item}
            if word["english"] not in self.words:
                self.words[word["english"]] = []

            word = {
                "type": "Feature",
                "geometry": self.data[item.code]["geometry"],
                "properties": {
                    **word,
                    "language": {"code": item.code, "name": item.language["name"]},
                },
            }
            self.words[word["properties"]["english"]].append(word)

        for item in data:
            item = SimpleNamespace(**item)

            log.info(f"Building repository for {item.code}: {item.language['name']}")
            item_path = os.path.join(self.repository, item.code)
            self.makepath(item_path)

            if item.speaker:
                item.speaker = transcode_and_copy_to_repository(item.speaker, item_path)
                # pp.pprint(item["properties"]["speaker"])

            if item.language:
                item.language = transcode_and_copy_to_repository(
                    item.language, item_path
                )
                # pp.pprint(item["properties"]["language"])

            if item.words:
                words = []
                for word in item.words:
                    word = transcode_and_copy_to_repository(word, item_path)
                    # pp.pprint(word)
                    words.append(word)
                    push_to_words(word, item)
                item.words = words

            self.languages[item.code] = self.data[item.code]

            with open(os.path.join(item_path, "index.json"), "w") as f:
                f.write(json.dumps(self.data[item.code]))

    def makepath(self, path):
        try:
            os.makedirs(path)
        except:
            pass

    def write_master_indices(self):
        languages = {}
        for (code, language) in self.data.items():
            language["properties"]["words"] = False
            languages[code] = language

        for (code, language) in self.languages.items():
            language["properties"]["words"] = True
            self.languages[code] = language
            languages[code] = language

        languages = [language for (code, language) in languages.items()]
        with open(f"{self.repository}/languages-with-data.json", "w") as f:
            f.write(json.dumps({"languages": self.languages}))

        with open(f"{self.repository}/languages.json", "w") as f:
            f.write(json.dumps({"languages": languages}))

        words = []
        for name in WORDS:
            try:
                word = self.words[name]
                m = hashlib.sha256()
                m.update(name.encode())
                index = m.hexdigest()
                words.append({"name": name, "index": f"{index}.json"})
                with open(f"{self.repository}/{index}.json", "w") as f:
                    f.write(json.dumps(word))
            except:
                pass

        with open(f"{self.repository}/words.json", "w") as f:
            f.write(json.dumps({"words": words}))

        with open(f"{self.repository}/errors.json", "w") as f:
            f.write(
                json.dumps(
                    {
                        "date": datetime.now().strftime("%Y-%m-%dT%H:%M:%SZ"),
                        "errors": self.errors,
                    }
                )
            )


if __name__ == "__main__":
    d = DataExtractor()
    d.extract()
